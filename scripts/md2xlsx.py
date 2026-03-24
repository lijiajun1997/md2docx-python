#!/usr/bin/env python3
"""
Markdown to Excel Converter
============================

将Markdown文件转换为Excel，支持：
- 多Sheet（一级标题创建新Sheet）
- 单元格内富文本格式（加粗、斜体、代码、删除线）
- 段落内容（标题、文本）
- 统一字体：微软雅黑

用法:
    python3 md2xlsx.py input.md output.xlsx

作者: CoPaw AI Assistant
版本: 1.3.1
"""

import sys
import re
import html as html_module
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


# 统一字体配置
DEFAULT_FONT_NAME = '微软雅黑'
DEFAULT_FONT_SIZE = 11


class RichTextParser:
    """富文本解析器"""
    
    @staticmethod
    def parse_to_rich_text(text, base_bold=False, base_italic=False):
        """解析文本并生成Excel富文本对象"""
        if not text:
            return ''
        
        # 处理HTML实体
        text = html_module.unescape(text)
        # 处理<br>换行
        text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
        # 移除其他HTML标签
        text = re.sub(r'</?[a-zA-Z][^>]*>', '', text)
        
        # 解析格式
        segments = RichTextParser._parse_formatting(text, base_bold, base_italic)
        
        # 如果只有一个普通片段，返回纯文本
        if len(segments) == 1 and not any(segments[0][1:]):
            return segments[0][0]
        
        # 创建富文本对象
        text_parts = []
        for seg_text, bold, italic, code, strike in segments:
            if not seg_text:
                continue
            
            # 跳过只有换行的片段（合并到其他部分）
            if seg_text == '\n':
                continue
            
            has_format = bold or italic or strike or code
            
            if has_format:
                font_kwargs = {}
                if bold:
                    font_kwargs['b'] = True
                if italic:
                    font_kwargs['i'] = True
                if strike:
                    font_kwargs['strike'] = True
                if code:
                    pass
                if not code:
                    font_kwargs['rFont'] = DEFAULT_FONT_NAME
                
                font = InlineFont(**font_kwargs)
                text_parts.append(TextBlock(font, seg_text))
            else:
                text_parts.append(seg_text)
        
        # 如果所有部分都被过滤了，返回空字符串
        if not text_parts:
            return ''
        
        return CellRichText(*text_parts)
    
    @staticmethod
    def _parse_formatting(text, base_bold=False, base_italic=False):
        """解析文本中的格式标记"""
        result = []
        i = 0
        n = len(text)
        
        while i < n:
            # 链接 [...](...)
            if text[i] == '[':
                end_bracket = text.find('](', i)
                if end_bracket != -1:
                    end_paren = text.find(')', end_bracket)
                    if end_paren != -1:
                        link_text = text[i+1:end_bracket]
                        inner = RichTextParser._parse_formatting(link_text, base_bold, base_italic)
                        result.extend(inner)
                        i = end_paren + 1
                        continue
            
            # 删除线 ~~...~~
            if text[i:i+2] == '~~':
                end = text.find('~~', i+2)
                if end != -1:
                    inner_text = text[i+2:end]
                    inner = RichTextParser._parse_formatting(inner_text, base_bold, base_italic)
                    for seg_text, bold, italic, code, _ in inner:
                        result.append((seg_text, bold, italic, code, True))
                    i = end + 2
                    continue
            
            # 加粗 **...** 或 __...__
            if text[i:i+2] in ('**', '__'):
                marker = text[i:i+2]
                end = text.find(marker, i+2)
                if end != -1:
                    inner_text = text[i+2:end]
                    inner = RichTextParser._parse_formatting(inner_text, True, base_italic)
                    result.extend(inner)
                    i = end + 2
                    continue
            
            # 斜体 *...* 或 _..._
            if text[i] in ('*', '_'):
                marker = text[i]
                if i+1 < n and text[i+1] == marker:
                    pass
                else:
                    end = i + 1
                    while end < n:
                        if text[end] == marker and (end+1 >= n or text[end+1] != marker):
                            break
                        end += 1
                    
                    if end < n:
                        inner_text = text[i+1:end]
                        inner = RichTextParser._parse_formatting(inner_text, base_bold, True)
                        result.extend(inner)
                        i = end + 1
                        continue
            
            # 代码 `...`
            if text[i] == '`':
                end = text.find('`', i+1)
                if end != -1:
                    code_text = text[i+1:end]
                    result.append((code_text, base_bold, base_italic, True, False))
                    i = end + 1
                    continue
            
            # 普通字符
            result.append((text[i], base_bold, base_italic, False, False))
            i += 1
        
        # 合并相邻的相同格式片段（包括换行符）
        merged = []
        for seg in result:
            seg_text, seg_bold, seg_italic, seg_code, seg_strike = seg
            
            # 如果当前片段只有换行符，合并到前一个片段
            if seg_text == '\n' and merged:
                prev = merged[-1]
                merged[-1] = (prev[0] + '\n', prev[1], prev[2], prev[3], prev[4])
            elif merged and merged[-1][1:] == seg[1:]:
                # 格式相同，合并文本
                merged[-1] = (merged[-1][0] + seg[0],) + seg[1:]
            else:
                merged.append(seg)
        
        return merged
    
    @staticmethod
    def clean_text(text):
        """清理所有Markdown标记"""
        if not text:
            return ''
        
        result = text
        result = html_module.unescape(result)
        result = re.sub(r'<br\s*/?>', '\n', result, flags=re.IGNORECASE)
        result = re.sub(r'</?[a-zA-Z][^>]*>', '', result)
        result = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', result)
        result = re.sub(r'\*\*([^*]+)\*\*', r'\1', result)
        result = re.sub(r'__([^_]+)__', r'\1', result)
        result = re.sub(r'(?<!\*)\*(?!\*)([^*]+)(?<!\*)\*(?!\*)', r'\1', result)
        result = re.sub(r'(?<!_)_(?!_)([^_]+)(?<!_)_(?!_)', r'\1', result)
        result = re.sub(r'`([^`]+)`', r'\1', result)
        result = re.sub(r'~~([^~]+)~~', r'\1', result)
        result = re.sub(r' +', ' ', result)
        
        return result.strip()


class MarkdownToExcel:
    """Markdown转Excel转换器"""
    
    # 样式定义
    HEADER_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # 浅灰色
    HEADER_FONT = Font(bold=True, name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
    NORMAL_FONT = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
    TITLE_FONT = Font(bold=True, name=DEFAULT_FONT_NAME, size=14)
    SUBTITLE_FONT = Font(bold=True, name=DEFAULT_FONT_NAME, size=12)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    WRAP_ALIGNMENT = Alignment(wrap_text=False, vertical='top')  # 默认非自动换行
    
    @staticmethod
    def _get_alignment(text, indent=False):
        """根据内容决定对齐方式"""
        has_newline = '\n' in str(text)
        wrap = has_newline  # 包含换行符才自动换行
        
        if indent:
            return Alignment(indent=1, wrap_text=wrap)
        return Alignment(wrap_text=wrap, vertical='top')
    
    @staticmethod
    def _contains_newline(value):
        """检查值是否包含换行符（支持字符串和CellRichText）"""
        if value is None:
            return False
        # CellRichText对象转换为字符串检查
        return '\n' in str(value)
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = None
        self.current_row = 1
        self.first_sheet_created = False
    
    def convert(self, md_content):
        """转换Markdown内容为Excel"""
        lines = md_content.split('\n')
        i = 0
        
        while i < len(lines):
            line = lines[i].rstrip()
            
            if not line:
                i += 1
                continue
            
            # 一级标题 - 创建新Sheet
            if line.startswith('# ') and not line.startswith('## '):
                self._create_sheet(line[2:].strip())
                i += 1
                continue
            
            # 二级标题
            if line.startswith('## '):
                self._add_subtitle(line[3:].strip())
                i += 1
                continue
            
            # 三级标题
            if line.startswith('### '):
                self._add_section(line[4:].strip())
                i += 1
                continue
            
            # 表格
            if line.startswith('|') and '|' in line[1:]:
                table_data, consumed = self._parse_table(lines, i)
                self._add_table(table_data)
                i += consumed
                continue
            
            # 无序列表
            if line.startswith(('- ', '* ', '• ')):
                items, consumed = self._parse_list(lines, i)
                self._add_list(items)
                i += consumed
                continue
            
            # 有序列表
            if re.match(r'^\d+\.\s', line):
                items, consumed = self._parse_list(lines, i, ordered=True)
                self._add_list(items, ordered=True)
                i += consumed
                continue
            
            # 引用块
            if line.startswith('>'):
                self._add_quote(line[1:].strip())
                i += 1
                continue
            
            # 分隔线
            if line in ('---', '***', '___'):
                self._add_separator()
                i += 1
                continue
            
            # 普通段落
            if line and not line.startswith('```'):
                self._add_paragraph(line)
                i += 1
                continue
            
            # 代码块跳过
            if line.startswith('```'):
                code_block, consumed = self._parse_code_block(lines, i)
                self._add_code_block(code_block)
                i += consumed
                continue
            
            i += 1
        
        # 删除默认Sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        return self.wb
    
    def _create_sheet(self, title):
        """创建新工作表"""
        if not self.first_sheet_created:
            self.ws = self.wb.active
            self.ws.title = title[:31]
            self.first_sheet_created = True
        else:
            self.ws = self.wb.create_sheet(title[:31])
        
        # 设置第一列宽度
        self.ws.column_dimensions['A'].width = 3.5
        self.current_row = 1
    
    def _add_subtitle(self, text):
        """添加二级标题"""
        if self.ws is None:
            return
        
        cell = self.ws.cell(row=self.current_row, column=1)
        cell.value = RichTextParser.clean_text(text)
        cell.font = self.TITLE_FONT
        self.current_row += 1
    
    def _add_section(self, text):
        """添加三级标题"""
        if self.ws is None:
            return
        
        cell = self.ws.cell(row=self.current_row, column=1)
        cell.value = RichTextParser.clean_text(text)
        cell.font = self.SUBTITLE_FONT
        self.current_row += 1
    
    def _add_paragraph(self, text):
        """添加段落"""
        if self.ws is None:
            return
        
        cell = self.ws.cell(row=self.current_row, column=2)
        cell.value = RichTextParser.parse_to_rich_text(text)
        cell.font = self.NORMAL_FONT
        cell.alignment = self._get_alignment(text)
        self.current_row += 1
    
    def _add_quote(self, text):
        """添加引用"""
        if self.ws is None:
            return
        
        cell = self.ws.cell(row=self.current_row, column=2)
        cell.value = f'"{RichTextParser.clean_text(text)}"'
        cell.font = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, italic=True)
        cell.alignment = self._get_alignment(text, indent=True)
        self.current_row += 1
    
    def _add_list(self, items, ordered=False):
        """添加列表"""
        if self.ws is None:
            return
        
        for idx, item in enumerate(items, 1):
            prefix = f'{idx}. ' if ordered else '• '
            cell = self.ws.cell(row=self.current_row, column=2)
            cell.value = prefix + RichTextParser.clean_text(item)
            cell.font = self.NORMAL_FONT
            cell.alignment = self._get_alignment(item, indent=True)
            self.current_row += 1
    
    def _add_separator(self):
        """添加分隔线"""
        if self.ws is None:
            return
        
        cell = self.ws.cell(row=self.current_row, column=2)
        cell.value = '─' * 40
        cell.font = self.NORMAL_FONT
        cell.alignment = self._get_alignment('─' * 40)
        self.current_row += 1
    
    def _add_code_block(self, code):
        """添加代码块"""
        if self.ws is None:
            return
        
        cell = self.ws.cell(row=self.current_row, column=2)
        cell.value = code
        cell.font = Font(name='Courier New', size=10)
        cell.alignment = self._get_alignment(code)
        self.current_row += 1
    
    def _parse_list(self, lines, start, ordered=False):
        """解析列表"""
        items = []
        i = start
        
        while i < len(lines):
            line = lines[i].rstrip()
            
            if ordered:
                if re.match(r'^\d+\.\s', line):
                    items.append(re.sub(r'^\d+\.\s+', '', line))
                    i += 1
                else:
                    break
            else:
                if line.startswith(('- ', '* ', '• ')):
                    items.append(line[2:].strip())
                    i += 1
                else:
                    break
        
        return items, i - start
    
    def _parse_table(self, lines, start):
        """解析表格"""
        table_data = []
        i = start
        
        while i < len(lines) and lines[i].startswith('|'):
            row = [cell.strip() for cell in lines[i].split('|')[1:-1]]
            
            # 跳过分隔行
            if all(set(cell) <= set('-:|') for cell in row):
                i += 1
                continue
            
            table_data.append(row)
            i += 1
        
        return table_data, i - start
    
    def _add_table(self, table_data):
        """添加表格"""
        if not table_data or self.ws is None:
            return
        
        num_cols = max(len(row) for row in table_data)
        start_row = self.current_row
        
        # 写入所有行
        for row_idx, row_data in enumerate(table_data):
            for col_idx, cell_data in enumerate(row_data):
                if col_idx >= num_cols:
                    continue
                    
                cell = self.ws.cell(row=self.current_row, column=col_idx + 2)  # 从第2列开始
                
                # 表头加粗
                is_header = (row_idx == 0)
                rich_value = RichTextParser.parse_to_rich_text(cell_data, base_bold=is_header)
                
                # 检查是否为纯文本数字
                if isinstance(rich_value, str):
                    parsed = self._try_parse_number(rich_value)
                    if parsed is not None:
                        cell.value = parsed
                        if isinstance(parsed, float) and abs(parsed) < 1:
                            cell.number_format = '0.00%'
                    else:
                        cell.value = rich_value
                else:
                    cell.value = rich_value
                
                cell.border = self.BORDER
                # 根据单元格最终值决定是否换行
                wrap = self._contains_newline(cell.value)
                cell.alignment = Alignment(wrap_text=wrap, vertical='top')
                
                if is_header:
                    cell.fill = self.HEADER_FILL
                    cell.font = self.HEADER_FONT
                else:
                    cell.font = self.NORMAL_FONT
            
            self.current_row += 1
        
        # 设置列宽
        for col_idx in range(num_cols):
            col_letter = get_column_letter(col_idx + 2)  # 从第2列开始
            self.ws.column_dimensions[col_letter].width = 20
        
        # 空一行
        self.current_row += 1
    
    def _parse_code_block(self, lines, start):
        """解析代码块"""
        code_lines = []
        i = start + 1
        
        while i < len(lines) and not lines[i].startswith('```'):
            code_lines.append(lines[i])
            i += 1
        
        return '\n'.join(code_lines), i - start + 1
    
    def _try_parse_number(self, text):
        """尝试解析数字"""
        if not text:
            return None
        
        clean = text.strip()
        
        # 百分比
        if clean.endswith('%'):
            try:
                return float(clean.rstrip('%').rstrip()) / 100
            except ValueError:
                return None
        
        # 千分位数字
        if ',' in clean:
            try:
                return float(clean.replace(',', ''))
            except ValueError:
                return None
        
        # 普通数字
        try:
            if '.' in clean:
                return float(clean)
            return int(clean)
        except ValueError:
            return None
    
    def save(self, output_path):
        """保存Excel文件"""
        self.wb.save(output_path)
        return output_path


def main():
    if len(sys.argv) < 3:
        print('用法: python3 md2xlsx.py input.md output.xlsx')
        sys.exit(1)
    
    input_path = Path(sys.argv[1])
    output_path = sys.argv[2]
    
    if not input_path.exists():
        print(f'错误: 文件不存在 - {input_path}')
        sys.exit(1)
    
    md_content = input_path.read_text(encoding='utf-8')
    
    converter = MarkdownToExcel()
    converter.convert(md_content)
    converter.save(output_path)
    
    print(f'转换完成: {output_path}')


if __name__ == '__main__':
    main()
